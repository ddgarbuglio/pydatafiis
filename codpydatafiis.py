# ================== IMPORTAÇÃO DE BIBLIOTECAS ==================

# Pandas para leitura, manipulação e escrita de dados (Excel/DataFrame)
import pandas as pd

# Importa todas as funções auxiliares do módulo personalizado (se houver)
from funcoes_fiis import *

# Playwright para abrir páginas web com JavaScript (scraping dinâmico)
from playwright.sync_api import sync_playwright

# BeautifulSoup para fazer o parsing (leitura) do HTML
from bs4 import BeautifulSoup

# OS para manipulação de arquivos (excluir, abrir, verificar existência)
import os

# Openpyxl para editar arquivos Excel já criados
from openpyxl import load_workbook

# ===============================================================


# ================== LEITURA DO ARQUIVO DE ENTRADA ==================

# Lê o arquivo Excel contendo os FIIs
# usecols=[0,1] garante que apenas as duas primeiras colunas sejam lidas
df = pd.read_excel(
    "AA_IN_DADOS_fiis_listagem.xlsx",
    usecols=[0, 1]
)

# Lista que armazenará os dados finais de cada FII
linhas_df_saida = []

# Número total de linhas do DataFrame (quantidade de FIIs)
nlinhas_df = df.shape[0]

# ===============================================================


# ================== FUNÇÃO PARA EXTRAIR INDICADORES ==================
# Função criada para buscar informações exibidas nos blocos "desc"
# da página (ex.: tipo de fundo, segmento, cotistas etc.)

def extrair_indicadores_desc(soup):
    dados = {}

    # Percorre todos os blocos com classe "desc"
    for bloco in soup.select("div.desc"):
        nome = bloco.select_one("span.name")
        valor = bloco.select_one("div.value span")

        # Garante que nome e valor existam
        if nome and valor:
            chave = nome.get_text(strip=True).upper()
            valor_txt = valor.get_text(strip=True)

            # Salva no dicionário
            dados[chave] = valor_txt

    return dados

# ===============================================================


# ================== LOOP PRINCIPAL ==================
# Percorre cada FII listado no Excel

for index, row in df.iterrows():

    # Quantidade de cotas do FII
    int_cotas = float(row["Ncotas"])

    # Exemplo de cálculo auxiliar (não usado depois)
    fim_cota_div = 0.12 * int_cotas

    # ================== IDENTIFICAÇÃO DO ATIVO ==================
    ativo_ticker = row["Ticker"]

    print("✔️ Ticker informado para busca: ", ativo_ticker.upper())
    print("⏳ Aguarde enquanto os dados são baixados... ")
    print(f"🧮 Fundo {index+1} de {nlinhas_df}")

    # =============================================================


    # ================== DEFINIÇÃO DE URL E ARQUIVOS ==================
    # URL do FII no site Investidor10
    URL = f"https://investidor10.com.br/fiis/{ativo_ticker}/"

    # Nome do arquivo HTML temporário
    ARQUIVO_SAIDA = f"pagina_{ativo_ticker.upper()}.html"
    arquivo_html = ARQUIVO_SAIDA

    # =============================================================


    # ================== PLAYWRIGHT: DOWNLOAD DO HTML ==================
    # Usa um navegador invisível para carregar a página completa (JS)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()

        # Função para bloquear recursos desnecessários
        # Isso acelera o carregamento
        def bloquear_recursos(route):
            if route.request.resource_type in ["image", "media", "font", "stylesheet"]:
                route.abort()
            else:
                route.continue_()

        # Aplica o bloqueio para todos os recursos
        page.route("**/*", bloquear_recursos)

        # Acessa a página do FII
        page.goto(URL, wait_until="domcontentloaded")

        # Aguarda 2 segundos para garantir renderização completa
        page.wait_for_timeout(2000)

        # Captura o HTML final renderizado
        html = page.content()

        # Salva o HTML localmente
        with open(ARQUIVO_SAIDA, "w", encoding="utf-8") as f:
            f.write(html)

        browser.close()

    print(f"HTML do ticker {ativo_ticker.upper()} salvo com sucesso!")

    # =============================================================


    # ================== LEITURA E PARSING DO HTML ==================
    with open(arquivo_html, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f, "lxml")

        # Extrai dados dos blocos "desc"
        indicadores_desc = extrair_indicadores_desc(soup)

    # Dicionário final com os indicadores do FII
    indicadores = {}

    # ================== 5.1 COTAÇÃO ==================
    cotacao_tag = soup.select_one("strong.livePrice")
    if cotacao_tag:
        indicadores["Cotação"] = cotacao_tag.get_text(strip=True)

    # ================== 5.2 P/VP ==================
    for card in soup.select("div._card"):
        titulo = card.select_one("span[title='P/VP']")
        if titulo:
            valor = card.select_one("div._card-body span")
            if valor:
                indicadores["P/VP"] = valor.get_text(strip=True)
            break

    # ================== 5.3 DY (12M) ==================
    for card in soup.select("div._card.dy"):
        titulo = card.select_one("span[title='Dividend Yield']")
        if titulo and "DY (12M)" in titulo.get_text():
            valor = card.select_one("div._card-body span")
            if valor:
                indicadores["DY (12M)"] = valor.get_text(strip=True)
            break

    # ================== 5.4 MÉDIA DOS ÚLTIMOS 6 DIVIDENDOS ==================
    valores_dividendos = []
    linhas = soup.select("#table-dividends-history tbody tr")

    # Pega os 6 dividendos mais recentes
    for linha in linhas[:6]:
        colunas = linha.find_all("td")

        if len(colunas) >= 4:
            valor_txt = colunas[3].get_text(strip=True)

            # Converte valor brasileiro para float
            valor = valor_txt.replace(".", "").replace(",", ".")

            try:
                valores_dividendos.append(float(valor))
            except ValueError:
                pass

    if len(valores_dividendos) == 6:
        media_dividendos = sum(valores_dividendos) / 6
        indicadores["Média últimos 6 dividendos"] = f"R$ {media_dividendos:.4f}"
    else:
        indicadores["Média últimos 6 dividendos"] = "N/D"

    # ================== 5.5 A 5.8 DADOS DESCRITIVOS ==================
    indicadores["Tipo de Fundo"] = indicadores_desc.get("TIPO DE FUNDO")
    indicadores["Segmento"] = indicadores_desc.get("SEGMENTO")
    indicadores["Valor Patrimonial"] = indicadores_desc.get("VALOR PATRIMONIAL")
    indicadores["Cotistas"] = indicadores_desc.get("NUMERO DE COTISTAS")

    # =============================================================


    # ================== IMPRESSÃO DOS RESULTADOS ==================
    print("******************************************************")
    print(f"📊 {ativo_ticker.upper()} – site Investidor10")
    print("Cotação:", indicadores.get("Cotação"))
    print("DY (12M):", indicadores.get("DY (12M)"))
    print("P/VP:", indicadores.get("P/VP"))
    print("Média últimos 6 dividendos:", indicadores.get("Média últimos 6 dividendos"))
    print("Tipo de Fundo:", indicadores.get("Tipo de Fundo"))
    print("Segmento:", indicadores.get("Segmento"))
    print("Valor Patrimonial:", indicadores.get("Valor Patrimonial"))
    print("Cotistas:", indicadores.get("Cotistas"))
    print("******************************************************")

    # Armazena os dados para o DataFrame final
    linhas_df_saida.append({
        "xTicker": row["Ticker"],
        "xNcotas": row["Ncotas"],
        "Cotacao": indicadores.get("Cotação"),
        "TotGasto": 0,
        "P/VP": indicadores.get("P/VP"),
        "MediaDivR$/Cota": indicadores.get("Média últimos 6 dividendos"),
        "TotDivid": 0,
        "DY12meses": indicadores.get("DY (12M)"),
        "TipodeFundo": indicadores.get("Tipo de Fundo"),
        "Segmento": indicadores.get("Segmento"),
        "ValorPatrimonial": indicadores.get("Valor Patrimonial"),
        "NCotistas": indicadores.get("Cotistas"),
        "GBRindex": 0
    })

    # ================== EXCLUSÃO DO HTML TEMPORÁRIO ==================
    if os.path.exists(arquivo_html):
        os.remove(arquivo_html)
        print(f"✔️ Arquivo {arquivo_html} removido com sucesso\n")
    else:
        print(f"❌ Arquivo {arquivo_html} não encontrado\n")

# ===============================================================


# ================== CRIAÇÃO DO DATAFRAME FINAL ==================
df_saida = pd.DataFrame(linhas_df_saida)

# Tratamento de strings monetárias
df_saida["Cotacao"] = df_saida["Cotacao"].str.replace("R$ ", "").str.replace(",", ".")
df_saida["P/VP"] = df_saida["P/VP"].str.replace(",", ".")
df_saida["MediaDivR$/Cota"] = df_saida["MediaDivR$/Cota"].str.replace("R$ ", "")

# Conversão para float
df_saida["Cotacao"] = df_saida["Cotacao"].astype(float)
df_saida["TotGasto"] = df_saida["TotGasto"].astype(float)
df_saida["P/VP"] = df_saida["P/VP"].astype(float)
df_saida["MediaDivR$/Cota"] = df_saida["MediaDivR$/Cota"].astype(float)
df_saida["TotDivid"] = df_saida["TotDivid"].astype(float)

# Cálculos finais
df_saida["TotGasto"] = df_saida["xNcotas"] * df_saida["Cotacao"]
df_saida["TotDivid"] = df_saida["xNcotas"] * df_saida["MediaDivR$/Cota"]
df_saida["GBRindex"] = (df_saida["TotDivid"] / df_saida["TotGasto"]).round(4)

# Totais gerais
total_gasto = df_saida["TotGasto"].sum()
total_div = df_saida["TotDivid"].sum()

# Salva o Excel
df_saida.to_excel("saida_arquivo.xlsx", sheet_name="Planilha_Fiis_atual", index=False)

# ================== EDIÇÃO DO EXCEL FINAL ==================
wb = load_workbook("saida_arquivo.xlsx")
ws = wb["Planilha_Fiis_atual"]

# Insere totais na planilha
ws["O1"] = "TOTAIS"
ws["O2"] = f"Total gasto em R$: {total_gasto}"
ws["O3"] = f"Acréscimo de dividendos em R$: {total_div}"

wb.save("saida_arquivo.xlsx")

# Abre o arquivo automaticamente
os.startfile("saida_arquivo.xlsx")

# ==============================================================
